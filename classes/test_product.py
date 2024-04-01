from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import time
import pytest
import openpyxl
import utils.product_constants as c

options = Options()
options.add_argument("start-maximized")
class Test_Sauce:
    def __init__(self):
        self.driver=webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        self.driver.maximize_window()
        self.driver.get(c.base_url)
       
    def test_valid_login(self):
        userNameInput = self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))
        userNameInput.send_keys(c.standartuser_id)
        passwordInput.send_keys(c.secret_sauce_id)
        loginButton = self.waitForElementVisible=((By.ID,"login-button"))
        loginButton.click()
    
    def test_products_filter(self):
        
       filter_dropdown = self.waitForElementVisible=((By.CLASS_NAME,c.product_sort_container_name))
       time.sleep(3)
       select = Select(filter_dropdown)
       time.sleep(3)
       select.select_by_visible_text(c.select_visible_text)
       time.sleep(3)

      
    def test_product_details(self): 
        
        product_link = self.waitForElementVisible=((By.CSS_SELECTOR, c.product_link_celector))
        product_link.click()
        product_name = self.waitForElementVisible=((By.CLASS_NAME,c.product_name))
        assert product_name != ""
        product_price = self.waitForElementVisible=((By.CLASS_NAME, c.product_price))
        assert product_price != ""
        product_description = self.waitForElementVisible=((By.CLASS_NAME,c.product_description))
        assert product_description != ""
        time.sleep(5)

    def getData():
        excelFile=openpyxl.load_workbook("data/test_add_to_cart.xlsx")
        selectedSheet=excelFile["Sayfa1"]

        totalRows=selectedSheet.max_row
        data=[]
        for i in range(1, totalRows+1):
               product_name=selectedSheet.cell(i,1).value
               product_price=selectedSheet.cell(i,2).value
               tupleData=(product_name, product_price)
               data.append(tupleData)

        return data   
       
    @pytest.mark.parametrize("product_name, product_price",getData()) 
    def test_add_to_cart(self, cart_product_name, cart_product_price):
        add_to_cart_button = self.waitForElementVisible=((By.ID,c.add_to_cart_button_id ))
        add_to_cart_button.click()
        cart_icon = self.waitForElementVisible=((By.CLASS_NAME,c.cart_icon_name ))
        cart_icon.click()
        cart_product_name = self.waitForElementVisible=((By.CLASS_NAME,c.cart_product_name ))
        assert cart_product_name != ""
        cart_product_price =self.waitForElementVisible=((By.CLASS_NAME, c.cart_product_price_name))
        assert cart_product_price != ""
        time.sleep(3)
        
    def waitForElementVisible(self,locator,timeout=5):
        userNameInput=self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput=self.waitForElementVisible=((By.ID,c.password_id))
        loginButton=self.waitForElementVisible=((By.ID,c.login_button_id))
        product_link=self.waitForElementVisible=((By.ID,c.product_link_celector))
        product_name = self.waitForElementVisible=((By.CLASS_NAME,c.product_name))
        product_price = self.waitForElementVisible=((By.CLASS_NAME, c.product_price))
        product_description = self.waitForElementVisible=((By.CLASS_NAME,c.product_description))
        add_to_cart_button = self.waitForElementVisible=((By.ID,c.add_to_cart_button_id ))
        cart_icon = self.waitForElementVisible=((By.CLASS_NAME,c.cart_icon_name ))
        cart_product_name = self.waitForElementVisible=((By.CLASS_NAME,c.cart_product_name ))
        cart_product_price =self.waitForElementVisible=((By.CLASS_NAME, c.cart_product_price_name))
        

testclass=Test_Sauce()
testclass.test_valid_login() 
testclass.test_products_filter()
testclass.test_product_details()
product_name = "Product Name"
product_price = "Product Price"
testclass.test_add_to_cart(product_name, product_price)




    

