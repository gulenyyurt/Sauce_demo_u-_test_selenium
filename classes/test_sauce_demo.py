from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec 
from selenium.webdriver.common.action_chains import ActionChains 
import utils.test_sauce_constants as c
import pytest
import openpyxl

class Test_Sauce: 
    def precondition(self):
        driver =webdriver.Chrome()
        driver.maximize_window()
        driver.get(c.BASE_URL)  
        return driver
    def setup_method(self): #pytest tarafından tanımlanan bir metod her test öncesi otomaik olarak çalıştırılır.
        self.driver =webdriver.Chrome()
        self.driver.maximize_window()
        self.driver.get(c.BASE_URL)  

    def teardown_method(self): #her test bitiminde çalışacak fonksiyon
        self.driver.quit()

    def test_invalid_login(self): 
        driver = self.precondition()
        userNameInput = driver.find_element(By.ID,c.username_id)
        passwordInput = driver.find_element(By.ID,c.password_id)
        sleep(2)
        userNameInput.send_keys("1")
        passwordInput.send_keys("1")
        sleep(2)
        loginButton = driver.find_element(By.ID,c.login_button_id)
        loginButton.click()

    def readInvalidDataFromExcel():
        excelFile=openpyxl.load_workbook("data/invalidd_login.xlsx")
        selectedSheet=excelFile["Sayfa1"]

        totalRows=selectedSheet.max_row
        data=[]
        for i in range(1, totalRows+1):
               username=selectedSheet.cell(i,1).value
               password=selectedSheet.cell(i,2).value
               tupleData=(username, password)
               data.append(tupleData)

        return data   

    @pytest.mark.parametrize("username,password",readInvalidDataFromExcel())
    def test_invalid_login(self,username,password): 
        userNameInput = self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))

        actions= ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()
        loginButton = self.waitForElementVisible=(By.ID,c.login_button_id)
        loginButton.click()
        sleep(2)
        errorMessage = self.waitForElementVisible=(By.XPATH,c.errorMessage_xpath)
        #print(errorMessage.text)
        testResult = errorMessage.text == c.errorMessage_text
        print(f"TEST SONUCU: {testResult}")
        errorMessage = self.waitForElementVisible=((By.XPATH,c.errorMessage_xpath))
        assert errorMessage.text == c.errorMessage_text


# Case 1     
# -Kullanıcı adı ve şifre alanları boş geçildiğinde uyarı mesajı olarak "Epic sadface: Username is required" gösterilmelidir.
    @pytest.mark.parametrize(c.username_id,c.password_id,[("","")])
    def test_null_value(self,username,password):
        userNameInput = self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))
        actions= ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,c.username_id)
        actions.send_keys_to_element(passwordInput,c.password_id)
        actions.perform()
        loginButton = self.waitForElementVisible=((By.ID,c.login_button_id))
        loginButton.click()
        sleep(2)
        expectedMessage = self.waitForElementVisible=(By.XPATH,c.expectedMessage_xpath)
        testResult = expectedMessage.text == c.expectedMessage_text
        print(f"Epic sadface: Username is required uyarı mesajı gösterilmiştir = {testResult}")
        errorMessage =self.waitForElementVisible=((By.XPATH,c.errorMessage_xpath))
        assert errorMessage.text == c.errorMessage_text


#Case 2
#-Sadece şifre alanı boş geçildiğinde uyarı mesajı olarak "Epic sadface: Password is required" gösterilmelidir.
    def test_null_password(self):
        driver = self.precondition()
        userNameInput = self.waitForElementVisible=(By.ID,c.username_id)
        userNameInput.send_keys(c.standartuser_id)
        sleep(2)
        loginButton = self.waitForElementVisible=(By.ID,c.login_button_id)
        loginButton.click()
    @pytest.mark.parametrize(c.username_id,c.password_id,[(c.standartuser_id)])
    def test_null_password(self, username, password):
        userNameInput =self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))

        actions= ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,c.username_id)
        actions.send_keys_to_element(passwordInput,c.password_idord)
        actions.perform()
        loginButton = self.waitForElementVisible=((By.ID,c.login_button_id))
        loginButton.click()
        sleep(2)
        expectedMessage = self.waitForElementVisible=(By.XPATH,c.expectedMessage_xpath)
        testResult = expectedMessage.text == c.expectedMessage_text
        print(f"Epic sadface: Password is required uyarı mesajı gösterilmiştir = {testResult}")
        errorMessage = self.waitForElementVisible=((By.XPATH,c.errorMessage_xpath))
        assert errorMessage.text == c.errorMessage_text

#case 3
#Kullanıcı adı "locked_out_user" şifre alanı "secret_sauce" gönderildiğinde "Epic sadface: Sorry, this user has been locked out." mesajı gösterilmelidir.
    def test_user_locked(self):
        driver = self.precondition()
        userNameInput = self.waitForElementVisible=(By.ID,c.username_id)
        userNameInput.send_keys(c.lockeduserout_id)
        sleep(2)
        passwordInput = self.waitForElementVisible=(By.ID,c.password_id)
        passwordInput.send_keys(c.secret_sauce_id)
        sleep(3)
        loginButton = self.waitForElementVisible=(By.ID,c.login_button_id)
    @pytest.mark.parametrize(c.username_id,c.password_id,[(c.lockeduserout_id,c.secret_sauce_id)])
    def test_user_locked(self,username,password):
        userNameInput =self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))

        actions= ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,c.username_id)
        actions.send_keys_to_element(passwordInput,c.password_id)
        actions.perform()
        loginButton = self.waitForElementVisible=((By.ID,c.login_button_id))
        loginButton.click()
        expectedMessage = self.waitForElementVisible=(By.XPATH,c.expectedMessage_xpath)
        testResult = expectedMessage.text == c.expectedMessage_text
        print(f"Epic sadface: Sorry, this user has been locked out. uyarı mesajı gösterilmiştir = {testResult}")
        errorMessage = self.waitForElementVisible=((By.XPATH,c.errorMessage_xpath))
        assert errorMessage.text == c.errorMessage_text

#case 4
#-Kullanıcı adı "standard_user" şifre "secret_sauce" gönderildiğinde kullanıcı "/inventory.html" sayfasına gönderilmelidir. Giriş yapıldıktan sonra kullanıcıya gösterilen ürün sayısı "6" adet olmalıdır.
    def test_products_list(self):
        driver = self.precondition()
        userNameInput = self.waitForElementVisible=(By.ID,c.username_id)
        userNameInput.send_keys(c.standartuser_id)
        passwordInput = self.waitForElementVisible=(By.ID,c.password_id)
        passwordInput.send_keys(c.secret_sauce_id)
        sleep(3)
        loginButton = self.waitForElementVisible=(By.ID,c.login_button_id)
        userNameInput = self.waitForElementVisible=((By.ID,c.username_id))
        passwordInput = self.waitForElementVisible=((By.ID,c.password_id))

        actions= ActionChains(self.driver)
        actions.send_keys_to_element(userNameInput,c.standartuser_id)
        actions.send_keys_to_element(passwordInput,c.secret_sauce_id)
        actions.perform()
        loginButton = self.waitForElementVisible=((By.ID,c.login_button_id))
        loginButton.click()
        sleep(2)
        driver.get(c.base_urll)
        sleep(2)
        productList = self.waitForElementVisible=(By.CLASS_NAME,c.inventory_item_id)
        print(f"Ürün sayısı {len(productList)} adettir.")
        sleep(3)

        baslik = self.waitForElementVisible=((By.XPATH,c.baslik_xpath))

        assert baslik.text == c.baslik_text

    def waitForElementVisible(self,locator,timeout=5):
        usernameInput=self.waitForElementVisible=((By.Id,c.username_id))
        passwordInput=self.waitForElementVisible=((By.Id,c.password_id))
        loginButton=self.waitForElementVisible=((By.Id,c.login_button_id))
        errorMessage=waitForElementVisible=((By.Id,c.errorMessage_text))

